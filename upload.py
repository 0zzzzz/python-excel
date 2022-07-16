import os
import openpyxl
import psycopg2


if os.name == 'nt':
    file_path = os.getcwd()
else:
    file_path = os.path.dirname(os.path.realpath(__file__))
workbook_path = openpyxl.load_workbook(f'{file_path}/названия точек.xlsm')
workbook_sheet = workbook_path.active
title = workbook_sheet.cell(row=1, column=2).value


def cell_value(index, column_num):
    """
    Функция для получения данных из таблицы excel
    """
    return workbook_sheet.cell(row=index, column=column_num).value


def get_data():
    """
    Функция агрегатор данных полученных из таблицы excel
    """
    rows_amount = workbook_sheet.max_row
    workbook_data = []
    for index in range(2, rows_amount+1):
        endpoint_id = cell_value(index, 1)
        endpoint_name = cell_value(index, 2)
        if endpoint_id or endpoint_name:
            workbook_data.append((endpoint_id, endpoint_name))
    return workbook_data


def db_connect():
    """
    Функция для подключения к базе данных
    """
    connection = None
    try:
        connection = psycopg2.connect(
            user='postgres',
            password='postgres',
            database='test_db',
            host='localhost',
            port='5432'
        )
        connection.autocommit = True
        cursor = connection.cursor()
        with connection.cursor() as cursor:
            cursor.execute(
                'CREATE TABLE IF NOT EXISTS endpoint_names (\
                    endpoint_id SERIAL PRIMARY KEY,\
                    endpoint_names  VARCHAR(100)'
                ');'
            )
            connection.commit()
        with connection.cursor() as cursor:
            workbook_data = get_data()
            for record in workbook_data:
                try:
                    cursor.execute(
                        'INSERT INTO endpoint_names\
                            (endpoint_id, endpoint_names) VALUES\
                            (%s, %s) ',
                        (record[0], record[1])
                    )
                except Exception:
                    cursor.execute(
                        'UPDATE endpoint_names SET endpoint_names=%s\
                            WHERE endpoint_id=%s',
                        (record[1], record[0])
                    )
                connection.commit()
            print('[Данные успешно загружены в базу]')
    except Exception as error:
        print('[Ошибка при работе с базой данных]', error)
    finally:
        if connection:
            connection.close()
            print('[Соединение с базой данных завершено]')


if __name__ == '__main__':
    db_connect()